#!/usr/bin/env python3
"""Parse the FAA TAF annual series (all facilities, both scenarios, MIN_YEAR onward) from
the zip already cached by test-faa-taf.py during stage 1.

Unlike test-faa-taf.py (which only prints a base-year/horizon-year summary for the
data-feasibility spike), this outputs one row per (airport, year, scenario) — the
grain airport_forecast_annual actually needs. Run after test-faa-taf.py at least once
so data/raw/APO100_TAF_Final_2025.zip exists.
"""
import io, json, os, sys, urllib.request, warnings, zipfile
warnings.filterwarnings("ignore")
import openpyxl

URL = "https://taf.faa.gov/Downloads/APO100_TAF_Final_2025.zip"
RAW = "data/raw/APO100_TAF_Final_2025.zip"
# All 3,319 TAF facilities, no hand-picked list. The series starts in 1976; scoring only
# uses 2014 onward (T-100 actuals do not go back further), so older years are dropped here
# rather than loaded and ignored -- 224k source rows down to a few thousand.
MIN_YEAR = 2014


def fetch():
    os.makedirs("data/raw", exist_ok=True)
    if not os.path.exists(RAW):
        print(f"downloading {URL} ...", file=sys.stderr)
        urllib.request.urlretrieve(URL, RAW)
    return zipfile.ZipFile(RAW)


def read(zf, name, value_cols):
    ws = openpyxl.load_workbook(io.BytesIO(zf.read(name)), read_only=True)["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = {}
    for r in rows:
        year = int(r[2])
        if year < MIN_YEAR:
            continue
        loc = str(r[0]).strip()
        out[(loc, int(r[1]), year)] = sum(int(r[c] or 0) for c in value_cols)
    return out


def read_airports(zf):
    """FAA's own airport dimension — LOCID, name, city, state, hub size. Used so the
    airports table is sourced from public data instead of hand-maintained."""
    ws = openpyxl.load_workbook(io.BytesIO(zf.read("Airports.xlsx")), read_only=True)["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    next(rows)
    out = {}
    for r in rows:
        loc = str(r[0]).strip()
        out[loc] = {
            "locid": loc,
            "name": str(r[2] or "").strip(),
            "city": str(r[3] or "").strip().title() or None,
            "state": str(r[4] or "").strip() or None,
            "hub_size": r[14],
        }
    return out


zf = fetch()
enpl = read(zf, "Enplanements.xlsx", [3, 4, 5, 6, 7])
ops = read(zf, "AirportsOperations.xlsx", [3, 4, 5, 6, 7, 8])

airports = read_airports(zf)
with open("data/out/faa-airports.json", "w") as f:
    json.dump(sorted(airports.values(), key=lambda a: a["locid"]), f, indent=2)
print(f"wrote {len(airports)} airports to data/out/faa-airports.json", file=sys.stderr)

keys = sorted(set(enpl) | set(ops))
rows = [
    {
        "airport": a,
        "year": y,
        "scenario": s,
        "enplanements": enpl.get((a, s, y)),
        "operations": ops.get((a, s, y)),
    }
    for (a, s, y) in keys
]

os.makedirs("data/out", exist_ok=True)
with open("data/out/faa-taf-annual.json", "w") as f:
    json.dump(rows, f, indent=2)
print(f"wrote {len(rows)} rows to data/out/faa-taf-annual.json", file=sys.stderr)
